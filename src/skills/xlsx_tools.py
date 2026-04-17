from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _cell_to_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("|", "\\|").replace("\n", "<br>")
    return s


def _trim_rows(rows: list[list[str]]) -> list[list[str]]:
    out = list(rows)
    while out and not any(c.strip() for c in out[-1]):
        out.pop()
    return out


def _sheet_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    max_cols = 0
    for row in ws.iter_rows(values_only=True):
        vals = [_cell_to_text(v) for v in row]
        while vals and vals[-1] == "":
            vals.pop()
        if len(vals) > max_cols:
            max_cols = len(vals)
        rows.append(vals)
    if max_cols == 0:
        return []
    normalized = [r + [""] * (max_cols - len(r)) for r in rows]
    return _trim_rows(normalized)


def convert_xlsx_file_to_markdown(xlsx_path: Path) -> Path:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        lines: list[str] = [f"# {xlsx_path.stem}", ""]

        for ws in wb.worksheets:
            lines.append(f"## {ws.title}")
            lines.append("")
            rows = _sheet_rows(ws)
            if not rows:
                lines.append("_（此 Sheet 无数据）_")
                lines.append("")
                continue

            header = rows[0]
            if not any(x.strip() for x in header):
                header = [f"col_{i+1}" for i in range(len(header))]
            header = [h if h else f"col_{i+1}" for i, h in enumerate(header)]
            data_rows = rows[1:] if len(rows) > 1 else []

            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            for r in data_rows:
                lines.append("| " + " | ".join(r) + " |")
            lines.append("")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    out = xlsx_path.with_suffix(".md")
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def convert_xlsx_path(path_str: str, recursive: bool = False) -> tuple[list[Path], list[str]]:
    p = Path(path_str).expanduser()
    if not p.is_absolute():
        p = (Path.cwd() / p).resolve()
    if not p.exists():
        return [], [f"路径不存在：{p}"]

    if p.is_file():
        if p.suffix.lower() != ".xlsx" or p.name.startswith("~$"):
            return [], [f"不是可处理的 xlsx 文件：{p}"]
        try:
            return [convert_xlsx_file_to_markdown(p)], []
        except Exception as e:  # noqa: BLE001
            return [], [f"转换失败：{p} | {e}"]

    globber = p.rglob if recursive else p.glob
    files = [f for f in globber("*.xlsx") if f.is_file() and not f.name.startswith("~$")]
    if not files:
        return [], [f"目录下未找到 xlsx：{p}"]

    outs: list[Path] = []
    errs: list[str] = []
    for f in sorted(files):
        try:
            outs.append(convert_xlsx_file_to_markdown(f))
        except Exception as e:  # noqa: BLE001
            errs.append(f"转换失败：{f} | {e}")
    return outs, errs
